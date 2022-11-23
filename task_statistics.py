import csv
import re
import itertools
import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Название вакансии
        salary (Salary): Оклад вакансии
        area_name (str): Регион вакансии
        published_at (datetime): Время публикации вакансии
    """

    def __init__(self, name: str, salary: "Salary", area_name: str, published_at: str):
        """Инициализирует объект Vacancy, выполняет конвертацию для поля published_at (в datetime).

        Args:
            name (str): Название вакансии
            salary (Salary): Оклад вакансии
            area_name (str): Регион вакансии
            published_at (str): Время публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = datetime.datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z")


class Salary:
    """Класс для представления зарплаты.

    Attributes:
        currency_to_rub (dict): (class attribute) Словарь для конвертации валют в рубль по курсу
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
    """

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from: int, salary_to: int, salary_currency: str):
        """Инициализирует объект Salary.

        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def get_rub_average(self):
        """Вычисляет среднюю зарплату из вилки оклада и переводит в рубли, если необходимо.

        Returns:
            float: Средняя зарплата в рублях
        """
        salary_average = (self.salary_from + self.salary_to) / 2
        if self.salary_currency == "RUR":
            return salary_average
        return salary_average * Salary.currency_to_rub[self.salary_currency]


class DataSet:
    """Класс для представления данных вакансий.

    Attributes:
        __file_name (str): Имя файла для обработки данных
        __list_naming (list): Названия столбцов таблицы
        reader (_reader): Объект чтения для чтения строк из файла
        vacancies (list): Список вакансий
        salary_by_year (dict): Словарь средней зарплаты по годам
        vacancies_count_by_year (dict): Словарь количества вакансий по годам
        selected_vacancy_salary_by_year (dict): Словарь средней зарплаты выбранной профессии по годам
        selected_vacancy_count_by_year (dict): Словарь количества выбранной профессии по годам
        vacancies_count_by_area (dict): Словарь количества вакансий по городам
        salary_by_area (dict): Словарь средней зарплаты по городам
        salary_by_area_appropriate (dict): Отсортированный по убыванию словарь средней зарплаты по городам, чья доля по
        количеству вакансий больше 1 процента от общего числа вакансий
        fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        fraction_by_area_appropriate (dict): Отсортированный по убыванию словарь доли вакансий от общего числа вакансий
        по городам, чья доля по количеству вакансий больше 1 процента от общего числа вакансий
        salary_by_area_sliced (dict): Отсортированный по убыванию словарь средней зарплаты по городам, содержащий только
        первые 10 элементов
        fraction_by_area_sliced (dict): Отсортированный по убыванию словарь доли вакансий по городам, содержащий только
        первые 10 элементов
    """

    def __init__(self, file_name: str):
        """Инициализирует объект DataSet.

        Args:
            file_name (str): Имя файла для обработки данных
        """
        self.__file_name = file_name
        self.__list_naming = None

    def get_clear_value(self, value):
        """Разделяет исходную строку по символу переноса строки, после чего очищает каждую отдельную строку от html тегов
        и лишних пробелов и возвращает строку, объединенную символами переноса строки.

        Args:
            value (str): Список строк для очистки

        Returns:
            str: Строка, состоящая из объединненых символами переноса строки строк, очищенных от html тегов и лишних пробелов
        """
        temp = value.split('\n')
        result = [self.get_clear_str(s) for s in temp]
        return '\n'.join(result)

    def get_clear_str(self, s):
        """Очищает строку от html тегов и лишних пробелов.

        Args:
            s (str): Строка для очистки

        Returns:
            str: Строка, очищенная от html тегов и лишних пробелов
        """
        str_without_tags = re.sub("<.*?>", '', s)
        str_without_spaces = ' '.join(str_without_tags.split())
        return str_without_spaces

    def is_empty_file(self):
        """Возвращает True, если файл пуст, либо False, если файл не пуст.

        Returns:
            bool: True или False
        """
        return False if self.__list_naming else True

    def csv_reader(self):
        """Открывает файл для чтения и получает названия столбцов csv файла."""
        vacancies = open(self.__file_name, 'r', encoding="utf-8-sig")
        self.reader = csv.reader(vacancies)
        self.__list_naming = next(self.reader)

    def csv_filter(self):
        """Считывает вакансии из файла, содержащие все необходимые данные, очищает их от лишних пробелов и html тегов
         и сохраняет их в список вакансий."""
        self.vacancies = []
        for line in self.reader:
            if len(line) == len(self.__list_naming) and '' not in line:
                name = self.get_clear_value(line[0])
                salary_from = int(float(self.get_clear_value(line[1])))
                salary_to = int(float(self.get_clear_value(line[2])))
                salary_currency = self.get_clear_value(line[3])
                salary = Salary(salary_from, salary_to, salary_currency)
                area_name = self.get_clear_value(line[4])
                published_at = self.get_clear_value(line[5])
                vacancy = Vacancy(name, salary, area_name, published_at)
                salary.rub_average = salary.get_rub_average()
                self.vacancies.append(vacancy)

    def get_statistics(self, selected_vacancy: str):
        """Производит расчет статистики по требуемой профессии.

        Args:
            selected_vacancy (str): Профессия, по которой требуется получить статистику
        """
        self.salary_by_year = {}
        self.vacancies_count_by_year = {}
        self.selected_vacancy_salary_by_year = {}
        self.selected_vacancy_count_by_year = {}
        self.vacancies_count_by_area = {}
        self.salary_by_area = {}
        self.salary_by_area_appropriate = {}
        self.fraction_by_area = {}
        self.fraction_by_area_appropriate = {}
        for vacancy in self.vacancies:
            year = int(vacancy.published_at.strftime("%Y"))
            salary = vacancy.salary.get_rub_average()
            if year not in self.salary_by_year:
                self.salary_by_year[year] = 0
                self.vacancies_count_by_year[year] = 0
            self.salary_by_year[year] += salary
            self.vacancies_count_by_year[year] += 1
            if selected_vacancy in vacancy.name and selected_vacancy != '':
                if year not in self.selected_vacancy_salary_by_year:
                    self.selected_vacancy_salary_by_year[year] = 0
                    self.selected_vacancy_count_by_year[year] = 0
                self.selected_vacancy_salary_by_year[year] += salary
                self.selected_vacancy_count_by_year[year] += 1
            area = vacancy.area_name
            if area not in self.salary_by_area:
                self.salary_by_area[area] = 0
            self.salary_by_area[area] += salary
            if area not in self.vacancies_count_by_area:
                self.vacancies_count_by_area[area] = 0
            self.vacancies_count_by_area[area] += 1
        for year in self.salary_by_year:
            self.salary_by_year[year] = int(self.salary_by_year[year] / self.vacancies_count_by_year[year])
        for area in self.salary_by_area:
            self.salary_by_area[area] = int(self.salary_by_area[area] / self.vacancies_count_by_area[area])
            self.fraction_by_area[area] = round(self.vacancies_count_by_area[area] / len(self.vacancies), 4)
            if int(self.fraction_by_area[area] * 100) >= 1:
                self.salary_by_area_appropriate[area] = self.salary_by_area[area]
                self.fraction_by_area_appropriate[area] = self.fraction_by_area[area]
        if selected_vacancy:
            for year in self.selected_vacancy_salary_by_year:
                if self.selected_vacancy_salary_by_year[year] != 0:
                    self.selected_vacancy_salary_by_year[year] = int(
                        self.selected_vacancy_salary_by_year[year] / self.selected_vacancy_count_by_year[year])
            if len(self.selected_vacancy_salary_by_year) == 0:
                self.selected_vacancy_salary_by_year[2022] = 0
                self.selected_vacancy_count_by_year[2022] = 0
        self.salary_by_area_appropriate = {k: v for k, v in
                                           sorted(self.salary_by_area_appropriate.items(), key=lambda item: item[1],
                                                  reverse=True)}
        self.fraction_by_area_appropriate = {k: v for k, v in
                                             sorted(self.fraction_by_area_appropriate.items(), key=lambda item: item[1],
                                                    reverse=True)}
        self.salary_by_area_sliced = dict(itertools.islice(self.salary_by_area_appropriate.items(), 10))
        self.fraction_by_area_sliced = dict(itertools.islice(self.fraction_by_area_appropriate.items(), 10))


class InputConnect:
    """Класс для ввода и вывода данных.

    Attributes:
        csv_file_name (str): Имя csv файла
        vacancy_name (str): Название профессии
    """

    def print_statistics(self, data_set: DataSet):
        """Печатает статистику на экран.

        Args:
            data_set (DataSet): Дата-сет статистики
        """
        print(f"Динамика уровня зарплат по годам: {data_set.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data_set.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data_set.selected_vacancy_salary_by_year}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {data_set.selected_vacancy_count_by_year}")
        print(
            f"Уровень зарплат по городам (в порядке убывания): {data_set.salary_by_area_sliced}")
        print(
            f"Доля вакансий по городам (в порядке убывания): {data_set.fraction_by_area_sliced}")

    def ask_user(self):
        """Получает необходимые данные ввода от пользователя."""
        self.csv_file_name = input("Введите название файла: ")
        # self.csv_file_name = "vacancies_by_year.csv"
        self.vacancy_name = input("Введите название профессии: ")

    def check_input(self):
        """Проверяет на корректность введенные пользователем данные.

        Returns:
            bool: True, если данные введены корректно, иначе False
        """
        if self.csv_file_name and self.vacancy_name:
            return True
        return False

    def __init__(self):
        """Инициализирует объект класса InputConnect и создает отчет по статистике при корректности введенных
        пользователем данных."""
        self.ask_user()
        if self.check_input():
            data_set = DataSet(self.csv_file_name)
            try:
                data_set.csv_reader()
            except StopIteration:
                print("Пустой файл")
            if not data_set.is_empty_file():
                data_set.csv_filter()
                if len(data_set.vacancies) == 0:
                    print("Нет данных")
                else:
                    data_set.get_statistics(self.vacancy_name)
                    Report.generate_excel(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                          data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                          data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                          self.vacancy_name)
                    Report.generate_image(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                          data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                          data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                          self.vacancy_name)
                    Report.generate_pdf(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                        data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                        data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                        self.vacancy_name)
                    self.print_statistics(data_set)


class Report:
    """Класс для представления отчёта.

    Attributes:
        excel_output_filename (str): (class attribute) Имя сгенерированного файла таблицы Excel
        plot_output_filename (str): (class attribute) Имя сгенерированного файла с графиками
        pdf_output_filename (str): (class attribute) Имя сгенерированного файла отчёта pdf
        bold_font (Font): (class attribute) Жирный шрифт для Excel
        thin_botder (Border): (class attribute) Тонкая рамка для Excel
    """
    excel_output_filename = "report.xlsx"
    plot_output_filename = "graph.png"
    pdf_output_filename = "report.pdf"
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    @staticmethod
    def adjust_columns_width(sheet: Worksheet):
        """Подгоняет ширину столбцов таблицы под их содержимое.

        Args:
            sheet (Worksheet): Лист, для которого требуется подогнать ширину столбцов
        """
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23

    @staticmethod
    def apply_thin_border(sheet: Worksheet, begin_row_i, begin_col_i, end_row_i, end_col_i):
        """Обводит выбранную область тонкой рамкой.

        Args:
            sheet (Worksheet): Лист, на котором требуется выполнить обводку
            begin_row_i (int): Индекс начальной строки
            begin_col_i (int): Индекс начального столбца
            end_row_i (int): Индекс конечной строки
            end_col_i (int): Индекс конечного столбца
        """
        for row in sheet.iter_rows(min_row=begin_row_i, min_col=begin_col_i, max_row=end_row_i, max_col=end_col_i):
            for cell in row:
                cell.border = Report.thin_border

    @staticmethod
    def generate_excel(salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                       vacancies_count_by_year: dict,
                       selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict,
                       vacancy_name: str):
        """Создаёт файл таблицы Excel на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        wb = Workbook()
        ws1 = wb.active
        Report.fill_year_sheet(ws1, salary_by_year, selected_vacancy_salary_by_year, vacancies_count_by_year,
                               selected_vacancy_count_by_year, vacancy_name)
        ws2 = wb.create_sheet()
        Report.fill_area_sheet(ws2, salary_by_area, fraction_by_area)
        wb.save(Report.excel_output_filename)

    @staticmethod
    def fill_year_sheet(sheet: Worksheet, salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                        vacancies_count_by_year: dict,
                        selected_vacancy_count_by_year: dict, vacancy_name: str):
        """Заполняет лист таблицы статистикой по годам.

        Args:
            sheet (Worksheet): Лист, который требуется заполнить статистикой по годам
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        sheet.title = "Статистика по годам"
        sheet["A1"] = "Год"
        sheet["A1"].font = Report.bold_font
        sheet["B1"] = "Средняя зарплата"
        sheet["B1"].font = Report.bold_font
        sheet["C1"] = f"Средняя зарплата - {vacancy_name}"
        sheet["C1"].font = Report.bold_font
        sheet["D1"] = "Количество вакансий"
        sheet["D1"].font = Report.bold_font
        sheet["E1"] = f"Количество вакансий - {vacancy_name}"
        sheet["E1"].font = Report.bold_font
        row_i, col_i = 2, 1
        for year in salary_by_year.keys():
            if year in selected_vacancy_salary_by_year:
                sheet.cell(row_i, col_i, year)
                sheet.cell(row_i, col_i + 1, salary_by_year[year])
                sheet.cell(row_i, col_i + 2, selected_vacancy_salary_by_year[year])
                sheet.cell(row_i, col_i + 3, vacancies_count_by_year[year])
                sheet.cell(row_i, col_i + 4, selected_vacancy_count_by_year[year])
                row_i += 1
        Report.adjust_columns_width(sheet)
        Report.apply_thin_border(sheet, 1, 1, row_i - 1, 5)

    @staticmethod
    def fill_area_sheet(sheet: Worksheet, salary_by_area: dict, fraction_by_area: dict):
        """Заполняет лист таблицы статистикой по городам.

        Args:
            sheet (Worksheet): Лист, который требуется заполнить статистикой по городам
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        """
        sheet.title = "Статистика по городам"
        sheet["A1"] = "Город"
        sheet["A1"].font = Report.bold_font
        sheet["B1"] = "Уровень зарплат"
        sheet["B1"].font = Report.bold_font
        sheet["D1"] = "Город"
        sheet["D1"].font = Report.bold_font
        sheet["E1"] = "Доля вакансий"
        sheet["E1"].font = Report.bold_font
        row_i, col_i = 2, 1
        for area in salary_by_area.keys():
            sheet.cell(row_i, col_i, area)
            sheet.cell(row_i, col_i + 1, salary_by_area[area])
            row_i += 1
        row_i, col_i = 2, 4
        for area in fraction_by_area.keys():
            sheet.cell(row_i, col_i, area)
            sheet.cell(row_i, col_i + 1, fraction_by_area[area])
            sheet.cell(row_i, col_i + 1).number_format = FORMAT_PERCENTAGE_00
            row_i += 1
        Report.adjust_columns_width(sheet)
        Report.apply_thin_border(sheet, 1, 1, row_i - 1, 2)
        Report.apply_thin_border(sheet, 1, 4, row_i - 1, 5)

    @staticmethod
    def generate_image(salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                       vacancies_count_by_year: dict,
                       selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict,
                       vacancy_name: str):
        """Создаёт файл графиков на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        fig, ((salary_by_year_axis, count_by_year_axis),
              (salary_by_area_axis, fraction_by_area_axis)) = plt.subplots(2, 2)
        Report.generate_salary_by_year_axis(salary_by_year_axis, salary_by_year, selected_vacancy_salary_by_year,
                                            vacancy_name)
        Report.generate_count_by_year_axis(count_by_year_axis, vacancies_count_by_year, selected_vacancy_count_by_year,
                                           vacancy_name)
        Report.generate_salary_by_area_axis(salary_by_area_axis, salary_by_area)
        Report.generate_fraction_by_area_axis(fraction_by_area_axis, fraction_by_area)
        fig.tight_layout()
        plt.savefig(Report.plot_output_filename, dpi=125)

    @staticmethod
    def generate_salary_by_year_axis(ax, salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                                     vacancy_name: str):
        """Создает диаграмму уровня зарплат (как общего, так и выбранной профессии) по годам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        average_salaries = [salary_by_year[year] for year in salary_by_year]
        selected_vacancy_average_salaries = [selected_vacancy_salary_by_year[year] for year in
                                             selected_vacancy_salary_by_year]
        labels = [year for year in salary_by_year]
        x = np.arange(len(labels))
        width = 0.35
        average_bar = ax.bar(x - width / 2, average_salaries, width, label="средняя з/п")
        selected_vacancy_bar = ax.bar(x + width / 2, selected_vacancy_average_salaries, width,
                                      label=f"з/п {vacancy_name}")
        ax.set_title("Уровень зарплат по годам")
        ax.set_xticks(x, labels, rotation=90)
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    @staticmethod
    def generate_count_by_year_axis(ax, vacancies_count_by_year: dict, selected_vacancy_count_by_year: dict,
                                    vacancy_name: str):
        """Создает диаграмму количества вакансий (как общего, так и выбранной профессии) по годам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        vacancies_count = [vacancies_count_by_year[year] for year in vacancies_count_by_year]
        selected_vacancy_count = [selected_vacancy_count_by_year[year] for year in
                                  selected_vacancy_count_by_year]
        labels = [year for year in vacancies_count_by_year]
        x = np.arange(len(labels))
        width = 0.35
        vacancies_count_bar = ax.bar(x - width / 2, vacancies_count, width, label="Количество вакансий")
        selected_vacancy_bar = ax.bar(x + width / 2, selected_vacancy_count, width,
                                      label=f"Количество вакансий\n{vacancy_name}")
        ax.set_title("Количество вакансий по годам")
        ax.set_xticks(x, labels, rotation=90)
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    @staticmethod
    def generate_salary_by_area_axis(ax, salary_by_area: dict):
        """Создает горизонтальную диаграмму уровня зарплат по городам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            salary_by_area (dict): Словарь средней зарплаты по городам
        """
        average_salary = [salary_by_area[area] for area in salary_by_area]
        area_labels = [Report.string_break_line(area) for area in salary_by_area]
        y = np.arange(len(area_labels))
        ax.barh(y, average_salary)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")
        ax.set_yticks(y, labels=area_labels, fontsize=6, horizontalalignment="right", verticalalignment="center")
        for label in (ax.get_xticklabels()):
            label.set_fontsize(8)
        ax.grid(True, axis='x')

    @staticmethod
    def generate_fraction_by_area_axis(ax, fraction_by_area: dict):
        """Создает круговую диаграмму количества вакансий по городам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        """
        fraction = [fraction_by_area[area] for area in fraction_by_area]
        others_fraction = 1 - sum(fraction)
        fraction = [others_fraction] + fraction
        fraction_labels = ["Другие"] + [area for area in fraction_by_area]
        ax.pie(fraction, labels=fraction_labels, textprops={"fontsize": 6})
        ax.set_title("Доля вакансий по городам")

    @staticmethod
    def generate_pdf(salary_by_year: dict, selected_vacancy_salary_by_year: dict, vacancies_count_by_year: dict,
                     selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict, vacancy_name: str):
        """Создает файл отчета pdf на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        abs_path_to_img = f"file:///{os.path.abspath(Report.plot_output_filename)}"
        pdf_template = template.render(
            {"vacancy_name": vacancy_name, "abs_path": abs_path_to_img, "salary_by_year": salary_by_year,
             "selected_vacancy_salary_by_year": selected_vacancy_salary_by_year,
             "vacancies_count_by_year": vacancies_count_by_year,
             "selected_vacancy_count_by_year": selected_vacancy_count_by_year,
             "salary_by_area": salary_by_area, "fraction_by_area": fraction_by_area})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Users\ek110\Desktop\Программы\Прочее ДЗ\Технологии программирования\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, Report.pdf_output_filename, configuration=config,
                           options={'enable-local-file-access': None})

    @staticmethod
    def string_break_line(s: str):
        """Возвращает исходную строку, если она не содержит пробелов и дефисов, иначе возвращает строку, разделенную
        переносами.

        Args:
            s (str): Строка

        Returns:
            str: Строка, разделенная переносами
        """
        str_splitted_by_space = s.split()
        str_splitted_by_hyphen = s.split('-')
        if len(str_splitted_by_space) == len(str_splitted_by_hyphen) == 1:
            return s
        elif len(str_splitted_by_space) > len(str_splitted_by_hyphen):
            return '\n'.join(str_splitted_by_space)
        else:
            return '-\n'.join(str_splitted_by_hyphen)


if __name__ == "__main__":
    app = InputConnect()
