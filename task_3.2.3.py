import cProfile
import concurrent.futures
import csv
import re
import itertools
import datetime
from os.path import isfile, join

import dateutil.tz
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
import matplotlib.pyplot as plt
import numpy as np
import multiprocessing as mp
from prettytable import prettytable

import csv_splitter


def profile(func):
    def wrapper(*args, **kwargs):
        datafn = func.__name__ + ".profile"
        prof = cProfile.Profile()
        retval = prof.runcall(func, *args, **kwargs)
        prof.dump_stats(datafn)
        return retval
    return wrapper

class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Название вакансии
        description (str): Описание вакансии
        key_skills (list): Список навыков, необходимых для вакансии
        experience_id (str): Необходимый опыт работы
        premium (bool): Тип вакансии (премиум или нет)
        employer_name (str): Имя компании
        salary (Salary): Оклад вакансии
        area_name (str): Регион вакансии
        published_at (datetime): Время публикации вакансии
    """

    def __init__(self, name: str, salary: "Salary", area_name: str, published_at: str, description: str = None, key_skills: list = None, experience_id: str = None, premium: str = None,
                 employer_name: str = None):
        """Инициализирует объект Vacancy, выполняет конвертацию для полей premium (в bool) и published_at (в datetime).

        Args:
            name (str): Название вакансии
            salary (Salary): Оклад вакансии
            area_name (str): Регион вакансии
            published_at (str): Время публикации вакансии
            description (str): Описание вакансии
            key_skills (list): Список навыков, необходимых для вакансии
            experience_id (str): Необходимый опыт работы
            premium (str): Тип вакансии (премиум или нет)
            employer_name (str): Имя компании

        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").name
        'Программист'
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").description
        'Описание вакансии'
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").key_skills
        ['Python', 'Git']
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").experience_id
        'between1And3'
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").premium
        False
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").employer_name
        'ООО Рога и Копыта'
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").salary.get_rub_average()
        25000.0
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").area_name
        'Екатеринбург'
        >>> Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта").published_at.isoformat()
        '2022-11-23T00:00:00+03:00'
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = Vacancy.convert_str_to_datetime_using_string_parsing(published_at)
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = True if premium == "True" else False
        self.employer_name = employer_name

    # @staticmethod
    # def convert_str_to_datetime_using_strptime(s: str) -> datetime:
    #     return datetime.datetime.strptime(s, "%Y-%m-%dT%H:%M:%S%z")

    @staticmethod
    def convert_str_to_datetime_using_string_parsing(s: str) -> datetime:
        year = int(s[:4])
        month = int(s[5:7])
        day = int(s[8:10])
        hour = int(s[11:13])
        minute = int(s[14:16])
        second = int(s[17:19])
        timezone_offset_hours_int = int(s[19:22])
        timezone_delta = datetime.timedelta(hours=timezone_offset_hours_int)
        if timezone_offset_hours_int < 0:
            timezone_delta *= -1
        timezone_offset = dateutil.tz.tzoffset(None, timezone_delta)
        return datetime.datetime(year, month, day, hour, minute, second, tzinfo=timezone_offset)

    # @staticmethod
    # def convert_str_to_datetime_using_dateutil_parser(s: str) -> datetime:
    #     return dateutil.parser.parse(s)

class Salary:
    """Класс для представления зарплаты.

    Attributes:
        currency_to_rub (dict): (class attribute) Словарь для конвертации валют в рубль по курсу
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
    """

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salary_from: int, salary_to: int, salary_currency: str, salary_gross: str = None):
        """Инициализирует объект Salary.

        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            salary_gross (str): Гросс
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_gross = salary_gross

    def get_rub_average(self):
        """Вычисляет среднюю зарплату из вилки оклада и переводит в рубли, если необходимо.

        Returns:
            float: Средняя зарплата в рублях

        >>> Salary(20000, 20000, "RUR", "Да").get_rub_average()
        20000.0
        >>> Salary(20000.0, 30000, "RUR", "Да").get_rub_average()
        25000.0
        >>> Salary(20000, 30000.0, "RUR", "Да").get_rub_average()
        25000.0
        >>> Salary(3000, 5000, "USD", "Да").get_rub_average()
        242640.0
        """
        salary_average = (self.salary_from + self.salary_to) / 2
        if self.salary_currency == "RUR":
            return salary_average
        return salary_average * Salary.currency_to_rub[self.salary_currency]


class DataSet:
    """Класс для представления данных вакансий.

    Attributes:
        __file_name (str): Имя файла для обработки данных
        __list_naming (list): Названия столбцов таблицы
        reader (_reader): Объект чтения для чтения строк из файла
        vacancies (list): Список вакансий
        vacancies_by_year (dict): Словарь вакансий по годам
        vacancies_length_before_filtering (int): Количество вакансий до фильтрации по параметру
                salary_by_year (dict): Словарь средней зарплаты по годам
        vacancies_count_by_year (dict): Словарь количества вакансий по годам
        selected_vacancy_salary_by_year (dict): Словарь средней зарплаты выбранной профессии по годам
        selected_vacancy_count_by_year (dict): Словарь количества выбранной профессии по годам
        vacancies_count_by_area (dict): Словарь количества вакансий по городам
        salary_by_area (dict): Словарь средней зарплаты по городам
        salary_by_area_appropriate (dict): Отсортированный по убыванию словарь средней зарплаты по городам, чья доля по
        количеству вакансий больше 1 процента от общего числа вакансий
        fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        fraction_by_area_appropriate (dict): Отсортированный по убыванию словарь доли вакансий от общего числа вакансий
        по городам, чья доля по количеству вакансий больше 1 процента от общего числа вакансий
        salary_by_area_sliced (dict): Отсортированный по убыванию словарь средней зарплаты по городам, содержащий только
        первые 10 элементов
        fraction_by_area_sliced (dict): Отсортированный по убыванию словарь доли вакансий по городам, содержащий только
        первые 10 элементов
    """

    def __init__(self, file_name: str):
        """Инициализирует объект DataSet.

        Args:
            file_name (str): Имя файла для обработки данных
        """
        self.__file_name = file_name
        self.__list_naming = None
        self.vacancies = []
        self.vacancies_by_year = {}
        self.salary_by_year = {}
        self.vacancies_count_by_year = {}
        self.selected_vacancy_salary_by_year = {}
        self.selected_vacancy_count_by_year = {}
        self.vacancies_count_by_area = {}
        self.salary_by_area = {}
        self.salary_by_area_appropriate = {}
        self.fraction_by_area = {}
        self.fraction_by_area_appropriate = {}

    @staticmethod
    def get_clear_value(value: str):
        """Разделяет исходную строку по символу переноса строки, после чего очищает каждую отдельную строку от html тегов
        и лишних пробелов и возвращает строку, объединенную символами переноса строки.

        Args:
            value (str): Список строк для очистки

        Returns:
            str: Строка, состоящая из объединненых символами переноса строки строк, очищенных от html тегов и лишних пробелов

        >>> DataSet.get_clear_value("<span>Текст</span>\\n<h1>Заголовок 1   уровня</h1>\\n1   2   3")
        'Текст\\nЗаголовок 1 уровня\\n1 2 3'
        >>> DataSet.get_clear_value("<h1>Без переноса</h1> строк")
        'Без переноса строк'
        >>> DataSet.get_clear_value("Без html-тегов\\nи лишних пробелов")
        'Без html-тегов\\nи лишних пробелов'
        """
        temp = value.split('\n')
        result = [DataSet.get_clear_str(s) for s in temp]
        return '\n'.join(result)

    @staticmethod
    def get_clear_str(s: str):
        """Очищает строку от html тегов и лишних пробелов.

        Args:
            s (str): Строка для очистки

        Returns:
            str: Строка, очищенная от html тегов и лишних пробелов

        >>> DataSet.get_clear_str("<p>Paragraph</p>")
        'Paragraph'
        >>> DataSet.get_clear_str("one two    three   4")
        'one two three 4'
        >>> DataSet.get_clear_str("<div><h1>Заголовок    1 уровня</h1></div>")
        'Заголовок 1 уровня'
        """
        str_without_tags = re.sub("<.*?>", '', s)
        str_without_spaces = ' '.join(str_without_tags.split())
        return str_without_spaces

    def is_empty_file(self):
        """Возвращает True, если файл пуст, либо False, если файл не пуст.

        Returns:
            bool: True или False
        """
        return False if self.__list_naming else True

    def csv_reader_all_years(self):
        """Открывает csv файлы, разделенные по годам, для чтения и заполняет список вакансий по годам."""
        csv_files_by_years_dir_path = "./splitted_csv/"
        years_filenames = [f for f in os.listdir(csv_files_by_years_dir_path) if isfile(join(csv_files_by_years_dir_path, f))]
        with mp.Manager() as manager:
            vacancies_by_year = manager.dict()
            years = [int(year_filename.split('.')[0]) for year_filename in years_filenames]
            paths = [csv_files_by_years_dir_path + year_filename for year_filename in years_filenames]
            with concurrent.futures.ProcessPoolExecutor(4) as executor:
                executor.map(DataSet.fill_vacancies_by_year, itertools.repeat(vacancies_by_year), paths, years)
            self.vacancies_by_year = dict(vacancies_by_year)

    @staticmethod
    def get_csv_reader_by_year(path_to_year_csv: str):
        """Возвращает csv reader и список названий столбцов для файла по указанному пути.

        Args:
            path_to_year_csv (str): Путь до csv файла по году.

        Returns:
            Csv reader и список названий столбцов для файла по указанному пути
        """
        vacancies_by_year = open(path_to_year_csv, 'r', encoding="utf-8-sig")
        reader_by_year = csv.reader(vacancies_by_year)
        list_naming = next(reader_by_year)
        return reader_by_year, list_naming

    @staticmethod
    def parse_vacancies_from_csv_by_year(vacancies_by_year: dict, reader_by_year, list_naming, year: int):
        """Заполняет список вакансий по указанному году, используя переданный csv reader.

        Args:
            vacancies_by_year (dict): Словарь, содержащий списки вакансий по годам
            reader_by_year: Сsv reader для указанного года
            list_naming (list): Список столбцов
            year (int): Год
        """
        vacancies = []
        for line in reader_by_year:
            if len(line) == len(list_naming) and '' not in line:
                name = line[0]
                salary_from = int(float(line[1]))
                salary_to = int(float(line[2]))
                salary_currency = line[3]
                salary = Salary(salary_from, salary_to, salary_currency)
                area_name = line[4]
                published_at = line[5]
                vacancy = Vacancy(name, salary, area_name, published_at)
                salary.rub_average = salary.get_rub_average()
                vacancies.append(vacancy)
        vacancies_by_year[year] = vacancies

    @staticmethod
    def fill_vacancies_by_year(vacancies_by_year: dict, path_to_year_csv: str, year: int):
        """Заполняет список вакансий по указанному csv файлу и году.

        Args:
            vacancies_by_year (dict): Словарь, содержащий списки вакансий по годам
            path_to_year_csv (str): Путь до csv файла, содержащий вакансии за указанный год
            year (int): Год
        """
        reader_by_year, list_naming = DataSet.get_csv_reader_by_year(path_to_year_csv)
        DataSet.parse_vacancies_from_csv_by_year(vacancies_by_year, reader_by_year, list_naming, year)

    @staticmethod
    def process_statistics_by_year(vacancies: list, salary_by_year: dict,
                                   vacancies_count_by_year: dict, selected_vacancy_salary_by_year: dict,
                                   selected_vacancy_count_by_year: dict, year: int, selected_vacancy: str):
        """Производит расчет статистики по требуемой профессии в указанный год.

        Args:
            vacancies (list): Список вакансий для обработки
            salary_by_year (dict): Словарь зарплат по годам
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_salary_by_year (dict): Словарь зарплат по годам для выбранной профессии
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            year (int): Год
            selected_vacancy (str): Профессия, по которой требуется получить статистику
        """
        salary_by_year_result = 0
        vacancies_count_by_year_result = 0
        selected_vacancy_salary_by_year_result = 0
        selected_vacancy_count_by_year_result = 0
        for vacancy in vacancies:
            salary = vacancy.salary.get_rub_average()
            salary_by_year_result += salary
            vacancies_count_by_year_result += 1
            if selected_vacancy in vacancy.name and selected_vacancy != '':
                selected_vacancy_salary_by_year_result += salary
                selected_vacancy_count_by_year_result += 1
        salary_by_year_result = int(salary_by_year_result / vacancies_count_by_year_result)
        salary_by_year[year] = salary_by_year_result
        vacancies_count_by_year[year] = vacancies_count_by_year_result
        if selected_vacancy:
            if selected_vacancy_salary_by_year_result != 0:
                selected_vacancy_salary_by_year_result = int(
                    selected_vacancy_salary_by_year_result / selected_vacancy_count_by_year_result)
        selected_vacancy_salary_by_year[year] = selected_vacancy_salary_by_year_result
        selected_vacancy_count_by_year[year] = selected_vacancy_count_by_year_result

    def process_statistics_all_years(self, selected_vacancy: str):
        """Производит расчет статистики для требуемой профессии по всем годам.

        Args:
            selected_vacancy (str): Профессия, по которой требуется получить статистику
        """
        with mp.Manager() as manager:
            salary_by_year = manager.dict()
            vacancies_count_by_year = manager.dict()
            selected_vacancy_salary_by_year = manager.dict()
            selected_vacancy_count_by_year = manager.dict()
            with concurrent.futures.ProcessPoolExecutor(4) as executor:
                vacancies = list(self.vacancies_by_year.values())
                years = list(self.vacancies_by_year.keys())
                executor.map(DataSet.process_statistics_by_year, vacancies,
                             itertools.repeat(salary_by_year),
                             itertools.repeat(vacancies_count_by_year),
                             itertools.repeat(selected_vacancy_salary_by_year),
                             itertools.repeat(selected_vacancy_count_by_year),
                             years,
                             itertools.repeat(selected_vacancy))
            self.salary_by_year = dict(salary_by_year)
            self.vacancies_count_by_year = dict(vacancies_count_by_year)
            self.selected_vacancy_salary_by_year = dict(selected_vacancy_salary_by_year)
            self.selected_vacancy_count_by_year = dict(selected_vacancy_count_by_year)
        vacancies_count = sum([len(self.vacancies_by_year[year]) for year in self.vacancies_by_year])
        for year in self.vacancies_by_year:
            for vacancy in self.vacancies_by_year[year]:
                salary = vacancy.salary.get_rub_average()
                area = vacancy.area_name
                if area not in self.salary_by_area:
                    self.salary_by_area[area] = 0
                self.salary_by_area[area] += salary
                if area not in self.vacancies_count_by_area:
                    self.vacancies_count_by_area[area] = 0
                self.vacancies_count_by_area[area] += 1
        for area in self.salary_by_area:
            self.salary_by_area[area] = int(self.salary_by_area[area] / self.vacancies_count_by_area[area])
            self.fraction_by_area[area] = round(self.vacancies_count_by_area[area] / vacancies_count, 4)
            if int(self.fraction_by_area[area] * 100) >= 1:
                self.salary_by_area_appropriate[area] = self.salary_by_area[area]
                self.fraction_by_area_appropriate[area] = self.fraction_by_area[area]
        self.salary_by_area_appropriate = {k: v for k, v in
                                           sorted(self.salary_by_area_appropriate.items(), key=lambda item: item[1],
                                                  reverse=True)}
        self.fraction_by_area_appropriate = {k: v for k, v in
                                             sorted(self.fraction_by_area_appropriate.items(), key=lambda item: item[1],
                                                    reverse=True)}
        self.salary_by_area_sliced = dict(itertools.islice(self.salary_by_area_appropriate.items(), 10))
        self.fraction_by_area_sliced = dict(itertools.islice(self.fraction_by_area_appropriate.items(), 10))

    def csv_reader(self):
        """Открывает файл для чтения и получает названия столбцов csv файла."""
        vacancies = open(self.__file_name, 'r', encoding="utf-8-sig")
        self.reader = csv.reader(vacancies)
        self.__list_naming = next(self.reader)

    def csv_filter_for_table(self):
        """Считывает вакансии из файла, содержащие все необходимые данные, очищает их от лишних пробелов и html тегов
         и сохраняет их в список вакансий, а также количество вакансий до фильтрации по параметру."""
        for line in self.reader:
            if len(line) == len(self.__list_naming) and '' not in line:
                name = DataSet.get_clear_value(line[0])
                description = DataSet.get_clear_value(line[1])
                key_skills = DataSet.get_clear_value(line[2]).split('\n')
                experience_id = DataSet.get_clear_value(line[3])
                premium = DataSet.get_clear_value(line[4])
                employer_name = DataSet.get_clear_value(line[5])
                salary_from = int(float(DataSet.get_clear_value(line[6])))
                salary_to = int(float(DataSet.get_clear_value(line[7])))
                salary_gross = DataSet.get_clear_value(line[8])
                salary_currency = DataSet.get_clear_value(line[9])
                salary = Salary(salary_from, salary_to, salary_currency, salary_gross)
                area_name = DataSet.get_clear_value(line[10])
                published_at = DataSet.get_clear_value(line[11])
                vacancy = Vacancy(name, salary, area_name, published_at, description, key_skills, experience_id, premium, employer_name)
                salary.rub_average = salary.get_rub_average()
                self.vacancies.append(vacancy)
        self.vacancies_length_before_filtering = len(self.vacancies)

    def formatter(self, filter_key=None, filter_value=None):
        """Выполняет фильтрацию списка вакансий, если задан параметр фильтрации и его значение.

        Args:
            filter_key (str): Параметр фильтрации
            filter_value (str): Значение параметра фильтрации
        """
        if filter_key and filter_value:
            self.vacancies = list(
                filter(lambda vacancy: DataSet.check_vacancy(vacancy, filter_key, filter_value), self.vacancies))

    def sorter(self, sorting_parameter=None, reverse_sort=False):
        """Выполняет сортировку списка вакансий, если задан параметр сортировки.

        Args:
            sorting_parameter (str): Параметр сортировки
            reverse_sort (bool): Флаг сортировки в обратном порядке (True, если требуется отсортировать в обратном порядке)
        """
        if sorting_parameter:
            if sorting_parameter == "Название":
                self.vacancies.sort(key=lambda v: v.name, reverse=reverse_sort)
            elif sorting_parameter == "Описание":
                self.vacancies.sort(key=lambda v: v.description, reverse=reverse_sort)
            elif sorting_parameter == "Навыки":
                self.vacancies.sort(key=lambda v: len(v.key_skills), reverse=reverse_sort)
            elif sorting_parameter == "Опыт работы":
                self.vacancies.sort(key=lambda v: self.convert_experience_to_int(v.experience_id), reverse=reverse_sort)
            elif sorting_parameter == "Премиум-вакансия":
                self.vacancies.sort(key=lambda v: v.premium, reverse=reverse_sort)
            elif sorting_parameter == "Компания":
                self.vacancies.sort(key=lambda v: v.employer_name, reverse=reverse_sort)
            elif sorting_parameter == "Оклад":
                self.vacancies.sort(key=lambda v: v.salary.rub_average, reverse=reverse_sort)
            elif sorting_parameter == "Название региона":
                self.vacancies.sort(key=lambda v: v.area_name, reverse=reverse_sort)
            elif sorting_parameter == "Дата публикации вакансии":
                self.vacancies.sort(key=lambda v: v.published_at, reverse=reverse_sort)

    def convert_experience_to_int(self, experience: str):
        """Преобразует требуемый опыт в число, используемое для дальнейшего сравнения.

        Args:
            experience (str): Опыт работы

        Returns:
            int: Опыт работы в виде числа
        """
        if experience == "noExperience":
            return 0
        elif experience == "between1And3":
            return 1
        elif experience == "between3And6":
            return 2
        else:
            return 3

    @staticmethod
    def check_vacancy(vacancy: "Vacancy", filter_key, filter_value):
        """Проверяет вакансию на соответствие значения параметра фильтрации.

        Args:
            vacancy (Vacancy): Вакансия
            filter_key (str): Параметр фильтрации
            filter_value (str): Значение параметра фильтрации

        Returns:
            bool: True, если вакансия соответствует фильтру, иначе False.

        >>> DataSet.check_vacancy(Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта"), "Название", "Аналитик")
        False
        >>> DataSet.check_vacancy(Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта"), "Название", "Программист")
        True
        >>> DataSet.check_vacancy(Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта"), "Опыт работы", "От 1 года до 3 лет")
        True
        >>> DataSet.check_vacancy(Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта"), "Навыки", "Python")
        True
        >>> DataSet.check_vacancy(Vacancy("Программист",Salary(20000, 30000, "RUR", "Да" ),"Екатеринбург","2022-11-23T00:00:00+0300","Описание вакансии",["Python", "Git"],"between1And3","Нет","ООО Рога и Копыта"), "Оклад", 21000)
        True
        """
        if filter_key == "Название":
            return vacancy.name == filter_value
        elif filter_key == "Описание":
            return vacancy.description == filter_value
        elif filter_key == "Опыт работы":
            return InputConnect.experience_naming[vacancy.experience_id] == filter_value
        elif filter_key == "Премиум-вакансия":
            if vacancy.premium:
                return True if filter_value == "Да" else False
            else:
                return True if filter_value == "Нет" else False
        elif filter_key == "Компания":
            return vacancy.employer_name == filter_value
        elif filter_key == "Оклад":
            salary_value = int(filter_value)
            salary_from = float(vacancy.salary.salary_from)
            salary_to = float(vacancy.salary.salary_to)
            if salary_from <= salary_value <= salary_to:
                return True
        elif filter_key == "Навыки":
            filter_skills = filter_value.split(", ")
            if all(filter_skill in vacancy.key_skills for filter_skill in filter_skills):
                return True
        elif filter_key == "Название региона":
            return vacancy.area_name == filter_value
        elif filter_key == "Дата публикации вакансии":
            return vacancy.published_at.strftime("%d.%m.%Y") == filter_value
        elif filter_key == "Идентификатор валюты оклада":
            return InputConnect.currency_naming[vacancy.salary.salary_currency] == filter_value
        return False


class InputConnect:
    """Класс для ввода и вывода данных.

    Attributes:
        experience_naming (dict): (class attribute) Словарь для перевода опыта с английского на русский
        salary_gross_naming (dict): (class attribute) Словарь для перевода гросс с английского на русский
        currency_naming (dict): (class attribute) Словарь для перевода идентификатора валюты оклада на русский
        valid_keys (list): (class attribute) Корректные названия параметров для фильтрации и сортировки
        columns (list): (class attribute) Столбцы для вывода таблицы
        csv_file_name (str): Имя csv файла
        filter_parameter (list): Параметр фильтрации и его значение
        sorting_parameter (list): Параметр сортировки и его значение
        is_sorting_parameter_reverse (bool): Флаг сортировки в обратном порядке
        vacancy_range (list): Диапазон вывода
        columns_to_print (list): Требуемые столбцы для вывода таблицы
    """

    experience_naming = {"noExperience": "Нет опыта", "between1And3": "От 1 года до 3 лет",
                         "between3And6": "От 3 до 6 лет", "moreThan6": "Более 6 лет"}
    salary_gross_naming = {"true": "Без вычета налогов", "false": "С вычетом налогов"}
    currency_naming = {"AZN": "Манаты",
                        "BYR": "Белорусские рубли",
                        "EUR": "Евро",
                        "GEL": "Грузинский лари",
                        "KGS": "Киргизский сом",
                        "KZT": "Тенге",
                        "RUR": "Рубли",
                        "UAH": "Гривны",
                        "USD": "Доллары",
                        "UZS": "Узбекский сум"}
    valid_keys = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия',
                  'Компания', 'Нижняя граница вилки оклада', 'Верхняя граница вилки оклада',
                  'Оклад указан до вычета налогов', 'Идентификатор валюты оклада', 'Название региона',
                  'Дата публикации вакансии', 'Оклад']
    columns = ["Название", "Описание", "Навыки", "Опыт работы", "Премиум-вакансия", "Компания", "Оклад",
               "Название региона", "Дата публикации вакансии"]

    def print_vacancies_table(self, vacancies, vacancy_from, vacancy_to, columns_to_print):
        """Печатает таблицу с вакансиями на экран.

        Args:
            vacancies (list): Список вакансий для печати
            vacancy_from (int): Нижняя граница диапазона вывода
            vacancy_to (int): Верхняя граница диапазона вывода
            columns_to_print (list): Требуемые столбцы для вывода таблицы
        """
        table = prettytable.PrettyTable()
        table.field_names = ['№'] + self.columns
        table.max_width = 20
        table.align = 'l'
        table.hrules = prettytable.ALL
        if vacancy_from > 0:
            vacancy_from -= 1
        if vacancy_to != len(vacancies):
            vacancy_to -= 1
        for i in range(len(vacancies)):
            vacancy = vacancies[i]
            result_row = [str(i + 1)]
            result_row.append(self.shorten_string(vacancy.name))
            result_row.append(self.shorten_string(vacancy.description))
            result_row.append(self.shorten_string('\n'.join(vacancy.key_skills)))
            result_row.append(InputConnect.experience_naming[vacancy.experience_id])
            result_row.append("Да" if vacancy.premium else "Нет")
            result_row.append(vacancy.employer_name)
            salary_gross = InputConnect.salary_gross_naming[vacancy.salary.salary_gross.lower()]
            salary_currency = InputConnect.currency_naming[vacancy.salary.salary_currency]
            salary = f"{format(vacancy.salary.salary_from, ',').replace(',', ' ')} - {format(vacancy.salary.salary_to, ',').replace(',', ' ')} ({salary_currency}) ({salary_gross})"
            result_row.append(salary)
            result_row.append(vacancy.area_name)
            result_row.append(vacancy.published_at.strftime("%d.%m.%Y"))
            table.add_row(result_row)
        print(table.get_string(start=vacancy_from, end=vacancy_to, fields=['№'] + list(columns_to_print)))

    def print_statistics(self, data_set: DataSet):
        """Печатает статистику на экран.

        Args:
            data_set (DataSet): Дата-сет статистики
        """
        print(f"Динамика уровня зарплат по годам: {data_set.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data_set.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data_set.selected_vacancy_salary_by_year}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {data_set.selected_vacancy_count_by_year}")
        print(
            f"Уровень зарплат по городам (в порядке убывания): {data_set.salary_by_area_sliced}")
        print(
            f"Доля вакансий по городам (в порядке убывания): {data_set.fraction_by_area_sliced}")

    def shorten_string(self, s):
        """Возвращает укороченную до 100-и символов строку, если её длина превышает 100 символов, иначе исходную строку.

        Args:
            s (str): Строка

        Returns:
            str: Укороченная  или исходная строка
        """
        if len(s) > 100:
            return f"{s[:100]}..."
        return s

    def ask_user(self):
        """Получает необходимые данные ввода от пользователя."""
        self.output_type = input()
        self.csv_file_name = input("Введите название файла: ")
        if self.output_type == "Вакансии":
            self.filter_parameter = input("Введите параметр фильтрации: ").split(": ")
            self.sorting_parameter = input("Введите параметр сортировки: ")
            self.is_sorting_parameter_reverse = input("Обратный порядок сортировки (Да / Нет): ")
            if self.is_sorting_parameter_reverse == "Да":
                self.is_sorting_parameter_reverse = True
            elif self.is_sorting_parameter_reverse == "Нет" or self.is_sorting_parameter_reverse == '':
                self.is_sorting_parameter_reverse = False
            self.vacancy_range = list(map(int, input("Введите диапазон вывода: ").split()))
            self.columns_to_print = input("Введите требуемые столбцы: ").split(', ')
            self.filter_key = self.filter_parameter[0]
        elif self.output_type == "Статистика":
            self.vacancy_name = input("Введите название профессии: ")

    def check_input(self):
        """Проверяет на корректность введенные пользователем данные.

        Returns:
            bool: True, если данные введены корректно, иначе False
        """
        if self.output_type == "Вакансии":
            if len(self.filter_parameter) == 1 and self.filter_key != '':
                print("Формат ввода некорректен")
                return False
            elif self.filter_key not in InputConnect.valid_keys and self.filter_key != '':
                print("Параметр поиска некорректен")
                return False
            elif self.sorting_parameter not in InputConnect.valid_keys and self.sorting_parameter != '':
                print("Параметр сортировки некорректен")
                return False
            elif self.is_sorting_parameter_reverse not in (True, False):
                print("Порядок сортировки задан некорректно")
                return False
            else:
                return True
        elif self.output_type == "Статистика":
            if self.csv_file_name and self.vacancy_name:
                return True
        return False

    @profile
    def __init__(self):
        """Инициализирует объект класса InputConnect и обрабатывает данные вакансий при корректности введенных
        пользователем данных."""
        self.ask_user()
        if self.check_input():
            data_set = DataSet(self.csv_file_name)
            try:
                data_set.csv_reader()
            except StopIteration:
                print("Пустой файл")
            if not data_set.is_empty_file():
                if self.output_type == "Вакансии":
                    data_set.csv_filter_for_table()
                    if len(self.filter_parameter) == 2:
                        filter_key, filter_value = self.filter_parameter
                        if filter_key in InputConnect.valid_keys:
                            data_set.formatter(filter_key, filter_value)
                    if len(self.vacancy_range) == 0:
                        vacancy_from = 0
                        vacancy_to = data_set.vacancies_length_before_filtering
                    elif len(self.vacancy_range) == 1:
                        vacancy_from = self.vacancy_range[0]
                        vacancy_to = data_set.vacancies_length_before_filtering
                    else:
                        vacancy_from, vacancy_to = self.vacancy_range
                    if len(self.columns_to_print) == 1 and self.columns_to_print[0] == '':
                        if len(data_set.vacancies) != 0:
                            self.columns_to_print = self.columns
                        else:
                            self.columns_to_print = []
                    if len(data_set.vacancies) == 0:
                        if data_set.vacancies_length_before_filtering != 0:
                            print("Ничего не найдено")
                        else:
                            print("Нет данных")
                    else:
                        data_set.sorter(self.sorting_parameter, self.is_sorting_parameter_reverse)
                        self.print_vacancies_table(data_set.vacancies, vacancy_from, vacancy_to, self.columns_to_print)
                elif self.output_type == "Статистика":
                    csv_splitter.split_csv_by_year(self.csv_file_name, "./splitted_csv/", "published_at")
                    data_set.csv_reader_all_years()
                    if len(data_set.vacancies_by_year) == 0:
                        print("Нет данных")
                    else:
                        data_set.process_statistics_all_years(self.vacancy_name)
                        Report.generate_excel(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                              data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                              data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                              self.vacancy_name)
                        Report.generate_image(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                              data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                              data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                              self.vacancy_name)
                        Report.generate_pdf(data_set.salary_by_year, data_set.selected_vacancy_salary_by_year,
                                            data_set.vacancies_count_by_year, data_set.selected_vacancy_count_by_year,
                                            data_set.salary_by_area_sliced, data_set.fraction_by_area_sliced,
                                            self.vacancy_name)
                        self.print_statistics(data_set)

class Report:
    """Класс для представления отчёта.

    Attributes:
        excel_output_filename (str): (class attribute) Имя сгенерированного файла таблицы Excel
        plot_output_filename (str): (class attribute) Имя сгенерированного файла с графиками
        pdf_output_filename (str): (class attribute) Имя сгенерированного файла отчёта pdf
        bold_font (Font): (class attribute) Жирный шрифт для Excel
        thin_botder (Border): (class attribute) Тонкая рамка для Excel
    """
    excel_output_filename = "report.xlsx"
    plot_output_filename = "graph.png"
    pdf_output_filename = "report.pdf"
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    @staticmethod
    def adjust_columns_width(sheet: Worksheet):
        """Подгоняет ширину столбцов таблицы под их содержимое.

        Args:
            sheet (Worksheet): Лист, для которого требуется подогнать ширину столбцов
        """
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23

    @staticmethod
    def apply_thin_border(sheet: Worksheet, begin_row_i, begin_col_i, end_row_i, end_col_i):
        """Обводит выбранную область тонкой рамкой.

        Args:
            sheet (Worksheet): Лист, на котором требуется выполнить обводку
            begin_row_i (int): Индекс начальной строки
            begin_col_i (int): Индекс начального столбца
            end_row_i (int): Индекс конечной строки
            end_col_i (int): Индекс конечного столбца
        """
        for row in sheet.iter_rows(min_row=begin_row_i, min_col=begin_col_i, max_row=end_row_i, max_col=end_col_i):
            for cell in row:
                cell.border = Report.thin_border

    @staticmethod
    def generate_excel(salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                       vacancies_count_by_year: dict,
                       selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict,
                       vacancy_name: str):
        """Создаёт файл таблицы Excel на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        wb = Workbook()
        ws1 = wb.active
        Report.fill_year_sheet(ws1, salary_by_year, selected_vacancy_salary_by_year, vacancies_count_by_year,
                               selected_vacancy_count_by_year, vacancy_name)
        ws2 = wb.create_sheet()
        Report.fill_area_sheet(ws2, salary_by_area, fraction_by_area)
        wb.save(Report.excel_output_filename)

    @staticmethod
    def fill_year_sheet(sheet: Worksheet, salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                        vacancies_count_by_year: dict,
                        selected_vacancy_count_by_year: dict, vacancy_name: str):
        """Заполняет лист таблицы статистикой по годам.

        Args:
            sheet (Worksheet): Лист, который требуется заполнить статистикой по годам
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        sheet.title = "Статистика по годам"
        sheet["A1"] = "Год"
        sheet["A1"].font = Report.bold_font
        sheet["B1"] = "Средняя зарплата"
        sheet["B1"].font = Report.bold_font
        sheet["C1"] = f"Средняя зарплата - {vacancy_name}"
        sheet["C1"].font = Report.bold_font
        sheet["D1"] = "Количество вакансий"
        sheet["D1"].font = Report.bold_font
        sheet["E1"] = f"Количество вакансий - {vacancy_name}"
        sheet["E1"].font = Report.bold_font
        row_i, col_i = 2, 1
        for year in salary_by_year.keys():
            if year in selected_vacancy_salary_by_year:
                sheet.cell(row_i, col_i, year)
                sheet.cell(row_i, col_i + 1, salary_by_year[year])
                sheet.cell(row_i, col_i + 2, selected_vacancy_salary_by_year[year])
                sheet.cell(row_i, col_i + 3, vacancies_count_by_year[year])
                sheet.cell(row_i, col_i + 4, selected_vacancy_count_by_year[year])
                row_i += 1
        Report.adjust_columns_width(sheet)
        Report.apply_thin_border(sheet, 1, 1, row_i - 1, 5)

    @staticmethod
    def fill_area_sheet(sheet: Worksheet, salary_by_area: dict, fraction_by_area: dict):
        """Заполняет лист таблицы статистикой по городам.

        Args:
            sheet (Worksheet): Лист, который требуется заполнить статистикой по городам
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        """
        sheet.title = "Статистика по городам"
        sheet["A1"] = "Город"
        sheet["A1"].font = Report.bold_font
        sheet["B1"] = "Уровень зарплат"
        sheet["B1"].font = Report.bold_font
        sheet["D1"] = "Город"
        sheet["D1"].font = Report.bold_font
        sheet["E1"] = "Доля вакансий"
        sheet["E1"].font = Report.bold_font
        row_i, col_i = 2, 1
        for area in salary_by_area.keys():
            sheet.cell(row_i, col_i, area)
            sheet.cell(row_i, col_i + 1, salary_by_area[area])
            row_i += 1
        row_i, col_i = 2, 4
        for area in fraction_by_area.keys():
            sheet.cell(row_i, col_i, area)
            sheet.cell(row_i, col_i + 1, fraction_by_area[area])
            sheet.cell(row_i, col_i + 1).number_format = FORMAT_PERCENTAGE_00
            row_i += 1
        Report.adjust_columns_width(sheet)
        Report.apply_thin_border(sheet, 1, 1, row_i - 1, 2)
        Report.apply_thin_border(sheet, 1, 4, row_i - 1, 5)

    @staticmethod
    def generate_image(salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                       vacancies_count_by_year: dict,
                       selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict,
                       vacancy_name: str):
        """Создаёт файл графиков на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        fig, ((salary_by_year_axis, count_by_year_axis),
              (salary_by_area_axis, fraction_by_area_axis)) = plt.subplots(2, 2)
        Report.generate_salary_by_year_axis(salary_by_year_axis, salary_by_year, selected_vacancy_salary_by_year,
                                            vacancy_name)
        Report.generate_count_by_year_axis(count_by_year_axis, vacancies_count_by_year, selected_vacancy_count_by_year,
                                           vacancy_name)
        Report.generate_salary_by_area_axis(salary_by_area_axis, salary_by_area)
        Report.generate_fraction_by_area_axis(fraction_by_area_axis, fraction_by_area)
        fig.tight_layout()
        plt.savefig(Report.plot_output_filename, dpi=125)

    @staticmethod
    def generate_salary_by_year_axis(ax, salary_by_year: dict, selected_vacancy_salary_by_year: dict,
                                     vacancy_name: str):
        """Создает диаграмму уровня зарплат (как общего, так и выбранной профессии) по годам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        average_salaries = [salary_by_year[year] for year in salary_by_year]
        selected_vacancy_average_salaries = [selected_vacancy_salary_by_year[year] for year in
                                             selected_vacancy_salary_by_year]
        labels = [year for year in salary_by_year]
        x = np.arange(len(labels))
        width = 0.35
        average_bar = ax.bar(x - width / 2, average_salaries, width, label="средняя з/п")
        selected_vacancy_bar = ax.bar(x + width / 2, selected_vacancy_average_salaries, width,
                                      label=f"з/п {vacancy_name}")
        ax.set_title("Уровень зарплат по годам")
        ax.set_xticks(x, labels, rotation=90)
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    @staticmethod
    def generate_count_by_year_axis(ax, vacancies_count_by_year: dict, selected_vacancy_count_by_year: dict,
                                    vacancy_name: str):
        """Создает диаграмму количества вакансий (как общего, так и выбранной профессии) по годам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            vacancy_name (str): Выбранная профессия
        """
        vacancies_count = [vacancies_count_by_year[year] for year in vacancies_count_by_year]
        selected_vacancy_count = [selected_vacancy_count_by_year[year] for year in
                                  selected_vacancy_count_by_year]
        labels = [year for year in vacancies_count_by_year]
        x = np.arange(len(labels))
        width = 0.35
        vacancies_count_bar = ax.bar(x - width / 2, vacancies_count, width, label="Количество вакансий")
        selected_vacancy_bar = ax.bar(x + width / 2, selected_vacancy_count, width,
                                      label=f"Количество вакансий\n{vacancy_name}")
        ax.set_title("Количество вакансий по годам")
        ax.set_xticks(x, labels, rotation=90)
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    @staticmethod
    def generate_salary_by_area_axis(ax, salary_by_area: dict):
        """Создает горизонтальную диаграмму уровня зарплат по городам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            salary_by_area (dict): Словарь средней зарплаты по городам
        """
        average_salary = [salary_by_area[area] for area in salary_by_area]
        area_labels = [Report.string_break_line(area) for area in salary_by_area]
        y = np.arange(len(area_labels))
        ax.barh(y, average_salary)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")
        ax.set_yticks(y, labels=area_labels, fontsize=6, horizontalalignment="right", verticalalignment="center")
        for label in (ax.get_xticklabels()):
            label.set_fontsize(8)
        ax.grid(True, axis='x')

    @staticmethod
    def generate_fraction_by_area_axis(ax, fraction_by_area: dict):
        """Создает круговую диаграмму количества вакансий по городам.

        Args:
            ax: Подграфик, на котором следует создать диаграмму
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
        """
        fraction = [fraction_by_area[area] for area in fraction_by_area]
        others_fraction = 1 - sum(fraction)
        fraction = [others_fraction] + fraction
        fraction_labels = ["Другие"] + [area for area in fraction_by_area]
        ax.pie(fraction, labels=fraction_labels, textprops={"fontsize": 6})
        ax.set_title("Доля вакансий по городам")

    @staticmethod
    def generate_pdf(salary_by_year: dict, selected_vacancy_salary_by_year: dict, vacancies_count_by_year: dict,
                     selected_vacancy_count_by_year: dict, salary_by_area: dict, fraction_by_area: dict, vacancy_name: str):
        """Создает файл отчета pdf на основе переданных данных.

        Args:
            salary_by_year (dict): Словарь средней зарплаты по годам
            selected_vacancy_salary_by_year (dict): Словарь средней зарплаты по годам для выбранной профессии
            vacancies_count_by_year (dict): Словарь количества вакансий по годам
            selected_vacancy_count_by_year (dict): Словарь количества вакансий по годам для выбранной профессии
            salary_by_area (dict): Словарь средней зарплаты по городам
            fraction_by_area (dict): Словарь доли вакансий от общего числа вакансий по городам
            vacancy_name (str): Выбранная профессия
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        abs_path_to_img = f"file:///{os.path.abspath(Report.plot_output_filename)}"
        pdf_template = template.render(
            {"vacancy_name": vacancy_name, "abs_path": abs_path_to_img, "salary_by_year": salary_by_year,
             "selected_vacancy_salary_by_year": selected_vacancy_salary_by_year,
             "vacancies_count_by_year": vacancies_count_by_year,
             "selected_vacancy_count_by_year": selected_vacancy_count_by_year,
             "salary_by_area": salary_by_area, "fraction_by_area": fraction_by_area})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Users\ek110\Desktop\Программы\Прочее ДЗ\Технологии программирования\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, Report.pdf_output_filename, configuration=config,
                           options={'enable-local-file-access': None})

    @staticmethod
    def string_break_line(s: str):
        """Возвращает исходную строку, если она не содержит пробелов и дефисов, иначе возвращает строку, разделенную
        переносами.

        Args:
            s (str): Строка

        Returns:
            str: Строка, разделенная переносами
        """
        str_splitted_by_space = s.split()
        str_splitted_by_hyphen = s.split('-')
        if len(str_splitted_by_space) == len(str_splitted_by_hyphen) == 1:
            return s
        elif len(str_splitted_by_space) > len(str_splitted_by_hyphen):
            return '\n'.join(str_splitted_by_space)
        else:
            return '-\n'.join(str_splitted_by_hyphen)

if __name__ == "__main__":
    app = InputConnect()